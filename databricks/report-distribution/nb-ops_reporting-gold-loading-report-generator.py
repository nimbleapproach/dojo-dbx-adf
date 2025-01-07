# Databricks notebook source
import json
import os

ENVIRONMENT = os.environ["__ENVIRONMENT__"]

current_notebook = dbutils.notebook.entry_point.getDbutils().notebook().getContext().notebookPath().get()
working_folder = '/Workspace' + current_notebook[0:current_notebook.rfind('/')]

# Loading config
report_config = "report.config.json"
with open(f"{working_folder}/{report_config}", "r") as file:
    data = json.load(file)

# COMMAND ----------

from datetime import datetime, timedelta

date_format = "%Y-%m-%d"

# Prompting report name and run date
try:
    REPORT_NAME = dbutils.widgets.get("wg_reportName")
except:
    dbutils.widgets.dropdown(name="wg_reportName", defaultValue=list(data.keys())[0], choices=sorted(data.keys()))
    REPORT_NAME = dbutils.widgets.get("wg_reportName")

try:
    REPORT_RUN_DATE = dbutils.widgets.get("wg_reportRunDate")
except:
    dbutils.widgets.text(name="wg_reportRunDate", defaultValue=datetime.strftime(datetime.now(), date_format))
    REPORT_RUN_DATE = dbutils.widgets.get("wg_reportRunDate")

# COMMAND ----------

# Reading config
report_config = data.get(REPORT_NAME)

report_params = report_config.get("report-params")
column_mapping = report_config.get("column-mapping")
report_totals = report_config.get("totals")

databricks_config = report_config.get("databricks-config")
databricks_predicate = databricks_config.get("predicate")
databricks_order_by = databricks_config.get("order-by")
databricks_object_layer = databricks_config.get("object-layer")
databricks_object_schema = databricks_config.get("object-schema")
databricks_object_name = databricks_config.get("object-name")

subscription_schedule = report_config.get("subscription").get("schedule")
from_date_cron = subscription_schedule.get("from-date-cron")
to_date_cron = subscription_schedule.get("to-date-cron")

subscription_recipients = report_config.get("subscription").get("recipients")
subscription_to = subscription_recipients.get("to")
subscription_cc = subscription_recipients.get("cc")
subscription_bcc = subscription_recipients.get("bcc")

# COMMAND ----------

# MAGIC %run ../library/nb-cron-utils

# COMMAND ----------

# calculating defaults for special parameters
report_param_defaults = {}

# report start date 
report_run_date_as_datetime = datetime.strptime(REPORT_RUN_DATE, date_format)
report_date_to = get_previous_date(report_run_date_as_datetime, to_date_cron) - timedelta(days=1)
report_param_defaults["to"] = datetime.strftime(report_date_to, date_format)

# report end date
report_date_from = get_previous_date(report_date_to, from_date_cron)
report_param_defaults["from"] = datetime.strftime(report_date_from, date_format)

# COMMAND ----------

input_params = {}

for report_param_name in report_params:
    widget_name = f"wg_reportParam_{report_param_name}"
    widget_default_value = report_param_defaults.get(report_param_name, "")

    try:
        report_param_value = dbutils.widgets.get(widget_name)
    except:
        dbutils.widgets.text(name=widget_name, defaultValue=widget_default_value)
        report_param_value = dbutils.widgets.get(widget_name)

    input_params[report_param_name] = report_param_value

# COMMAND ----------

from pyspark.sql.functions import col, sum
from pyspark.sql import Window

column_mapping = dict((v, k) for k, v in column_mapping.items())
report_columns = column_mapping.values()
view_columns_in_scope = [col(c) for c in column_mapping.keys()]

databricks_param_mapping = {}
for param_name, param_value in input_params.items():
    databricks_param_mapping[f"@{param_name}"] = param_value
# print(databricks_param_mapping)

def substitute_params(str, substitutions):
    result = str
    for bind_var, replacement in substitutions.items():
        result = result.replace(bind_var, replacement)
    return result


effective_databricks_predicate = substitute_params(databricks_predicate, databricks_param_mapping)
# print(effective_databricks_predicate)

# debug_predicate="1=1"
sorting_elements = [sorting_element.strip().split(" ") for sorting_element in databricks_order_by.split(",")]
sorting_columns = [sorting_tokens[0].strip() for sorting_tokens in sorting_elements]
sorting_ascending = [sorting_tokens[1].strip().lower() == "asc" if len(sorting_tokens) > 1 else True for sorting_tokens in sorting_elements]

report_totals = report_totals if report_totals else {}
total_columns = {}
for total in report_totals:
    total_column_base_name = total.get("base_column_name")
    total_column_display_name = total.get("display_name")
    total_column_aggregation_type = total.get("aggregation_type")

    if total_column_aggregation_type == "sum":
        total_column_aggregation_func = sum
    else:
        raise Exception(f"Unsupported aggregation type: {total_column_aggregation_type}")

    total_columns[total_column_display_name] = total_column_aggregation_func(col(total_column_base_name)).over(
        Window.partitionBy())

databricks_table = (spark.read
                    .table(
    f"{databricks_object_layer}_{ENVIRONMENT}.{databricks_object_schema}.{databricks_object_name}")
                    .filter(effective_databricks_predicate)
                    # .where(debug_predicate)
                    .select(view_columns_in_scope)
                    .sort(sorting_columns, ascending=sorting_ascending)
                    )

if (len(total_columns) > 0):
    databricks_table = databricks_table.withColumns(total_columns)

databricks_table = databricks_table.withColumnsRenamed(column_mapping)

report_data_types = dict(databricks_table.dtypes)
# print(report_data_types)

dataset = databricks_table.collect()

# COMMAND ----------

!pip install openpyxl --quiet

from copy import copy
from datetime import datetime
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Loading report template
template = "template.xlsx"
wb = load_workbook(working_folder + "/" + template)
sheet = wb.active
totals_row_count = len(report_totals)

# Template-specific values
report_name_cell_name = "A1"
date_range_cell_name = "A3"

total_template_first_cell_col = "A"
total_template_last_cell_col = "C"
total_template_cell_row = 4

header_template_cell_col = "A"
header_template_cell_row = 4

sheet_header_row_count = total_template_cell_row + totals_row_count

data_template_cell_col = "A"
data_template_cell_row = header_template_cell_row + 1

date_range_input_format = "%Y-%m-%d"
date_range_print_format = "%d/%m/%Y"

cell_date_format = '[$-en-US,1]dd/mm/yyyy'
cell_number_format = '[$-en-US,1]#,##0.00'

splitter_row_height = 5
min_column_width = 10
max_column_width = 70
column_width_factor = 1.2

logo_row = 0
logo_row_offset = 60000 # in EMUs (EMU is equivalent to 1/360,000 of a centimeter)
logo_length = 3 # in cells
logo_height = 1 # in cells

attachment_mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

header_template_cell_name = header_template_cell_col + str(header_template_cell_row)
header_template_cell = sheet[header_template_cell_name]
header_cell_font = copy(header_template_cell.font)
header_cell_fill = copy(header_template_cell.fill)
header_cell_border = copy(header_template_cell.border)

data_template_cell_name = data_template_cell_col + str(data_template_cell_row)
data_template_cell = sheet[data_template_cell_name]
data_cell_font = copy(data_template_cell.font)
data_cell_border = copy(data_template_cell.border)

# Positioning logo
image_ = sheet._images[0]

anchor = TwoCellAnchor()
anchor._from.row = logo_row
anchor._from.rowOff = logo_row_offset
anchor._from.col = len(report_columns) + 1
anchor.to.row = logo_row + logo_height
anchor.to.rowOff = logo_row_offset
anchor.to.col = len(report_columns) + 1 + logo_length

image_.anchor = anchor

# Populating sheet name
sheet.title = REPORT_NAME

# Populating report name header
sheet[report_name_cell_name] = REPORT_NAME

# Populating date range header
formatted_date_from = datetime.strftime(datetime.strptime(input_params["from"], date_range_input_format),
                                        date_range_print_format)
formatted_date_to = datetime.strftime(datetime.strptime(input_params["to"], date_range_input_format),
                                      date_range_print_format)

date_range_header_text = f"Report Period From {formatted_date_from} To {formatted_date_to}"
sheet[date_range_cell_name] = date_range_header_text

# Populating totals
current_row = total_template_cell_row

for total in report_totals:
    total_column_base_name = total.get("base_column_name")
    total_column_display_name = total.get("display_name")
    total_column_aggregation_type = total.get("aggregation_type")

    total_value = dataset[0][total_column_display_name] if len(dataset) > 0 else 0

    sheet.insert_rows(current_row)
    total_first_cell = total_template_first_cell_col + str(current_row)
    total_last_cell = total_template_last_cell_col + str(current_row)
    sheet.merge_cells(f"{total_first_cell}:{total_last_cell}")

    total_cell = sheet[total_first_cell]
    total_cell.value = f"{total_column_display_name} {total_value:,.2f}"
    total_cell.alignment = Alignment(horizontal="left", vertical="center")
    total_cell.font = header_cell_font

    current_row = current_row + 1

# Populating data column headers
for i, col in enumerate(report_columns):
    c = sheet.cell(row=sheet_header_row_count, column=i + 1)
    c.value = col
    c.font = header_cell_font
    c.fill = header_cell_fill
    c.border = header_cell_border

# Populating data cells
for i, data_row in enumerate(dataset):
    for j, data_col in enumerate(report_columns):
        # print(data_row.asDict())
        c = sheet.cell(row=sheet_header_row_count + i + 1, column=j + 1)
        c.value = data_row[data_col]
        c.border = data_cell_border
        c.font = data_cell_font

        data_type = report_data_types[data_col]
        if (data_type == "float") or (data_type == "double") or (data_type.startswith("decimal")):
            c.number_format = cell_number_format
        elif (data_type == "date") or (data_type == "timestamp") or (data_type == "timestampNTZ"):
            c.number_format = cell_date_format
        
# Auto-sizing columns
for column_cells in sheet.columns:
    max_length = max(0 if cell.value is None else len(str(cell.value)) for cell in column_cells[sheet_header_row_count - 1:])    
    column_letter = get_column_letter(column_cells[0].column)
    sheet.column_dimensions[column_letter].width = max(min_column_width, min(max_column_width, max_length * column_width_factor))

# Freezing header
header_row_count = total_template_cell_row + totals_row_count
sheet.insert_rows(header_row_count)
sheet.row_dimensions[header_row_count].height = splitter_row_height
sheet.freeze_panes = total_template_first_cell_col + str(header_row_count + 2)

# Saving result
report_folder = working_folder + "/reports"
os.makedirs(report_folder, exist_ok=True)

timestamp = datetime.strftime(datetime.today(), "%Y%m%d%H%M%S")
report_file_name = f"{REPORT_NAME}-{timestamp}.xlsx"
report_file_path = report_folder + "/" + report_file_name
wb.save(report_file_path)

# COMMAND ----------

# MAGIC %run ../library/nb-email-utils

# COMMAND ----------

import getpass

sender_email = "test@test.org"

smtp_login = sender_email
smtp_password = getpass.getpass("SMTP password")
smtp_host = "smtp.test.co.uk"
smtp_port = 587

# COMMAND ----------

subject = f"{REPORT_NAME} subscription"
body = f"{REPORT_NAME} extract for the period from {input_params['from']} to {input_params['to']}"

sendEmail(subject, sender_email, subscription_to, subscription_cc, subscription_bcc, body, {report_file_path : attachment_mime_type},
          smtp_host, smtp_port, smtp_login, smtp_password)
